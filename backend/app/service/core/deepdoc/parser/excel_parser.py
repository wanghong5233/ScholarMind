#
#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

from openpyxl import load_workbook
import sys
from io import BytesIO

from service.core.rag.nlp import find_codec


class RAGFlowExcelParser:
    def html(self, fnm, chunk_rows=256):
        try:
            if isinstance(fnm, str):
                wb = load_workbook(fnm)
            else:
                # 验证二进制数据不为空
                if not fnm:
                    raise ValueError("Excel 文件内容为空")
                
                # 验证文件格式（检查文件头）
                if not fnm.startswith(b'PK'):
                    raise ValueError("不是有效的 XLSX 文件格式，可能是 XLS 文件或文件已损坏")
                
                wb = load_workbook(BytesIO(fnm))
        except Exception as e:
            # 提供更友好的错误信息
            if "File is not a zip file" in str(e):
                raise ValueError("文件格式错误：XLSX 文件应该是 ZIP 格式，请检查文件是否完整或格式是否正确")
            elif "BadZipFile" in str(e):
                raise ValueError("文件损坏：无法解析 XLSX 文件，请重新上传")
            else:
                raise ValueError(f"Excel 文件解析失败: {str(e)}")

        tb_chunks = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(
                    rows[1 + chunk_i * chunk_rows : 1 + (chunk_i + 1) * chunk_rows]
                ):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def __call__(self, fnm):
        try:
            if isinstance(fnm, str):
                wb = load_workbook(fnm)
            else:
                # 验证二进制数据不为空
                if not fnm:
                    raise ValueError("Excel 文件内容为空")
                
                # 验证文件格式（检查文件头）
                if not fnm.startswith(b'PK'):
                    raise ValueError("不是有效的 XLSX 文件格式，可能是 XLS 文件或文件已损坏")
                
                wb = load_workbook(BytesIO(fnm))
        except Exception as e:
            # 提供更友好的错误信息
            if "File is not a zip file" in str(e):
                raise ValueError("文件格式错误：XLSX 文件应该是 ZIP 格式，请检查文件是否完整或格式是否正确")
            elif "BadZipFile" in str(e):
                raise ValueError("文件损坏：无法解析 XLSX 文件，请重新上传")
            else:
                raise ValueError(f"Excel 文件解析失败: {str(e)}")
        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            ti = list(rows[0])
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                line = "; ".join(fields)
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = load_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
